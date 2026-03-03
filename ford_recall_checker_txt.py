import openpyxl
from openpyxl.styles import Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime
import re
import os

def setup_driver(browser='chrome'):
    """Setup Chrome or Edge driver with appropriate options"""
    if browser == 'edge':
        from selenium.webdriver.edge.options import Options as EdgeOptions
        edge_options = EdgeOptions()
        edge_options.add_argument('--no-sandbox')
        edge_options.add_argument('--disable-dev-shm-usage')
        edge_options.add_argument('--disable-blink-features=AutomationControlled')
        edge_options.add_argument('--log-level=3')  # Suppress error messages
        edge_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        driver = webdriver.Edge(options=edge_options)
    else:
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--log-level=3')  # Suppress error messages
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        driver = webdriver.Chrome(options=chrome_options)
    return driver

def close_survey_popup(driver):
    """Close any survey popup that appears"""
    # List of possible close button selectors for various Ford popups
    close_selectors = [
        'button[data-aut="button-close"]',
        'button[aria-label="Close"]',
        'button[aria-label="close"]',
        '.modal-close',
        'button.close',
        # Ford survey modal X button (top-right)
        'div[class*="QSIWebResponsive"] button',
        'div[class*="modal"] button[class*="close"]',
        'div[class*="survey"] button',
    ]

    for selector in close_selectors:
        try:
            close_btns = driver.find_elements(By.CSS_SELECTOR, selector)
            for btn in close_btns:
                if btn.is_displayed():
                    try:
                        btn.click()
                        time.sleep(0.5)
                        print("  → Closed popup")
                        return True
                    except:
                        pass
        except:
            pass

    # Also try pressing Escape to close modal
    try:
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        time.sleep(0.3)
    except:
        pass

    return False

def wait_for_overlays_to_clear(driver, timeout=3):
    """Wait for any overlays/modals to disappear before interacting with elements"""
    try:
        # Wait for common overlay classes to not be present/visible
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, '.modal-overlay, .popup-overlay, [class*="overlay"]:not([class*="no-overlay"])'))
        )
    except:
        pass  # No overlay found or timeout - continue anyway

    # Also try to close any survey/feedback popups
    close_survey_popup(driver)

def setup_debug_log(output_dir):
    """Create a debug log file for tracking navigation issues"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(output_dir, f"DEBUG_LOG_{timestamp}.txt")
    return open(log_path, 'w')

def debug_log(log_file, vin, message):
    """Write debug message to log file"""
    if log_file:
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_file.write(f"[{timestamp}] VIN {vin}: {message}\n")
        log_file.flush()  # Ensure it's written immediately

def check_ford_recall(driver, vin, log_file=None):
    """
    Check Ford recall status for a given VIN using Selenium
    Returns: dict with hasRecall and recalls list (each recall has number, description, and remedy_available)
    """
    url = "https://www.ford.com/support/recalls-details/"
    
    try:
        wait = WebDriverWait(driver, 15)

        # Always navigate to fresh page for each VIN to avoid stale state issues
        print(f"  → Loading page...")
        driver.get(url)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
        debug_log(log_file, vin, f"URL after load: {driver.current_url}")

        print(f"  → Entering VIN...")
        # Wait for overlays to clear and close any popups
        wait_for_overlays_to_clear(driver)

        # Find and click the input field with retry logic
        max_retries = 3
        for attempt in range(max_retries):
            try:
                close_survey_popup(driver)
                vin_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", vin_input)
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", vin_input)  # Use JS click - more reliable
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"  → Retry {attempt + 1}: click blocked, waiting...")
                    debug_log(log_file, vin, f"Click attempt {attempt + 1} failed: {str(e)[:50]}")
                    time.sleep(1)
                    driver.execute_script("window.scrollTo(0, 0);")
                    wait_for_overlays_to_clear(driver)
                else:
                    raise
        time.sleep(0.3)

        # DEBUG: Log the VIN we're trying to enter
        debug_log(log_file, vin, f"Attempting to enter VIN: {vin}")
        print(f"  → DEBUG: Entering VIN: {vin}")

        # Get current value before clearing
        before_value = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value BEFORE clear: '{before_value}'")

        # Clear field using CTRL+A then DELETE
        vin_input.send_keys(Keys.CONTROL + "a")
        time.sleep(0.1)
        vin_input.send_keys(Keys.DELETE)
        time.sleep(0.1)

        # Check it's clear
        after_clear = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value AFTER clear: '{after_clear}'")

        # Type the VIN character by character
        for char in vin:
            vin_input.send_keys(char)
        time.sleep(0.2)

        # Verify VIN was entered correctly
        entered_value = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value AFTER typing: '{entered_value}'")
        print(f"  → DEBUG: Field now contains: {entered_value}")

        if entered_value != vin:
            debug_log(log_file, vin, f"WARNING: VIN mismatch! Field has: '{entered_value}', Expected: '{vin}'")
            print(f"  → VIN MISMATCH! Expected: {vin}, Got: {entered_value}")
            print(f"  → Retrying with fresh page...")
            # Retry with completely fresh page
            driver.get(url)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
            time.sleep(0.5)
            vin_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
            driver.execute_script("arguments[0].click();", vin_input)
            time.sleep(0.2)
            for char in vin:
                vin_input.send_keys(char)
            time.sleep(0.2)
            entered_value = vin_input.get_attribute('value')
            debug_log(log_file, vin, f"After retry, field value: '{entered_value}'")
            print(f"  → After retry, field contains: {entered_value}")

        debug_log(log_file, vin, f"VIN entered, about to submit")
        time.sleep(0.5)  # Wait before submit to let form validate
        vin_input.send_keys(Keys.RETURN)
        debug_log(log_file, vin, f"RETURN pressed")

        # Wait for results to load
        print(f"  → Waiting for results...")
        time.sleep(2)  # Short wait first

        # Check for error messages on page
        try:
            page_text = driver.find_element(By.TAG_NAME, "body").text
            if "unexpected error" in page_text.lower() or "error occurred" in page_text.lower():
                debug_log(log_file, vin, f"ERROR MESSAGE DETECTED on page!")
                # Save screenshot
                screenshot_path = os.path.join(os.path.dirname(log_file.name), f"ERROR_{vin}_{datetime.now().strftime('%H%M%S')}.png")
                driver.save_screenshot(screenshot_path)
                debug_log(log_file, vin, f"Screenshot saved: {screenshot_path}")
                # Try to find specific error element
                error_elements = driver.find_elements(By.CSS_SELECTOR, '[class*="error"], [class*="Error"], [role="alert"]')
                for el in error_elements[:3]:  # Log first 3 error elements
                    debug_log(log_file, vin, f"Error element text: {el.text[:100] if el.text else 'empty'}")
        except Exception as e:
            debug_log(log_file, vin, f"Error checking for errors: {str(e)[:50]}")

        time.sleep(3)  # Rest of wait time

        # Debug: Log URL after submission
        debug_log(log_file, vin, f"URL after submit: {driver.current_url}")

        # Check if we got redirected away from recalls page
        if '/recalls-details/' not in driver.current_url:
            debug_log(log_file, vin, f"Redirect detected, navigating back...")
            print(f"  → Redirect detected, navigating back...")
            driver.get(url)
            time.sleep(3)  # Wait for cached results
            debug_log(log_file, vin, f"URL after redirect: {driver.current_url}")
        else:
            # No redirect - wait for results to load on current page
            # Look for "no recalls" text OR Safety Recalls section
            debug_log(log_file, vin, "No redirect, waiting for results...")
            print(f"  → Waiting for results to load...")
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: 'no recalls' in d.find_element(By.TAG_NAME, "body").text.lower() or
                              d.find_elements(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]') or
                              'there are no' in d.find_element(By.TAG_NAME, "body").text.lower()
                )
                debug_log(log_file, vin, "Results detected on page")
            except:
                debug_log(log_file, vin, "Timeout waiting for results, continuing anyway")
                time.sleep(2)  # Fallback wait

        # Get page text after VIN submission
        body_text = driver.find_element(By.TAG_NAME, "body").text
        debug_log(log_file, vin, f"Page title: {driver.title}")

        print(f"  → Analyzing results...")
        
        # Check for "no recalls" message
        if 'no recalls' in body_text.lower() or 'there are no recalls' in body_text.lower():
            return {
                'hasRecall': False,
                'recalls': []
            }
        
        # Look for recall information
        recall_info = {
            'hasRecall': False,
            'recalls': []  # List of dicts: [{'number': '25S72', 'description': 'Desc...', 'remedy_available': True/False}]
        }
        
        try:
            # Find the Safety Recalls section specifically (not Customer Satisfaction Programs)
            safety_header = driver.find_elements(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]')

            if not safety_header:
                print(f"  → No Safety Recalls section found")
                return {
                    'hasRecall': False,
                    'recalls': []
                }

            # Find the tablist div that follows the safety recalls header (sibling)
            # The structure is: header div -> sibling div[role="tablist"] containing buttons
            try:
                # Get parent of header, then find the tablist within that parent
                parent = safety_header[0].find_element(By.XPATH, '..')
                tablist = parent.find_element(By.CSS_SELECTOR, '[role="tablist"]')
                recall_buttons = tablist.find_elements(By.CSS_SELECTOR, 'button[data-testid^="button-"][role="tab"]')
            except:
                # Fallback: try to find buttons near the safety header
                recall_buttons = []

            if not recall_buttons:
                print(f"  → No recall buttons found in Safety Recalls section")
                return {
                    'hasRecall': False,
                    'recalls': []
                }

            print(f"  → Found {len(recall_buttons)} recall button(s)")

            for idx in range(len(recall_buttons)):
                try:
                    # Re-find the buttons each time (DOM may change after clicking)
                    safety_header = driver.find_element(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]')
                    parent = safety_header.find_element(By.XPATH, '..')
                    tablist = parent.find_element(By.CSS_SELECTOR, '[role="tablist"]')
                    buttons = tablist.find_elements(By.CSS_SELECTOR, 'button[data-testid^="button-"][role="tab"]')

                    if idx >= len(buttons):
                        break

                    button = buttons[idx]

                    # Extract recall number from data-testid
                    testid = button.get_attribute('data-testid') or ''
                    recall_number = testid.replace('button-', '') if testid.startswith('button-') else None

                    if not recall_number:
                        continue

                    # Check for survey popup before interacting
                    close_survey_popup(driver)

                    # Scroll button into view
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                    time.sleep(0.5)

                    # Get description using JavaScript (more reliable)
                    recall_description = driver.execute_script("""
                        var p = arguments[0].querySelector('p');
                        return p ? p.innerText : '';
                    """, button) or 'See Ford website for details'

                    print(f"    [{idx+1}] Recall: {recall_number} - {recall_description[:50]}...")

                    # Click the button using JavaScript (more reliable than .click())
                    driver.execute_script("arguments[0].click();", button)
                    time.sleep(1.5)  # Wait for panel to load

                    # Extract full Campaign number (manufacturer/NHTSA) from the detail panel
                    try:
                        panel_id = f"content-panel-{recall_number}"
                        panel = driver.find_element(By.ID, panel_id)

                        # Search for Campaign section and get the value
                        campaign_text = driver.execute_script("""
                            var panel = arguments[0];
                            var sections = panel.querySelectorAll('section');
                            for (var i = 0; i < sections.length; i++) {
                                var label = sections[i].querySelector('p');
                                if (label && label.innerText.includes('Campaign')) {
                                    var valueP = sections[i].querySelectorAll('p')[1];
                                    if (valueP) return valueP.innerText.trim();
                                }
                            }
                            return null;
                        """, panel)

                        if campaign_text:
                            recall_number = campaign_text  # e.g., "23S55/23V675"
                            print(f"        → Campaign: {recall_number}")
                    except Exception as e:
                        print(f"        → Could not get full campaign #, using NHTSA: {recall_number}")

                    # Check for remedy status - simple page text search
                    page_text = driver.find_element(By.TAG_NAME, "body").text

                    if "Recall service not available right now" in page_text:
                        remedy_available = False
                        print(f"        → Remedy: NOT Available")
                    elif "Schedule your free recall service with a dealer" in page_text:
                        remedy_available = True
                        print(f"        → Remedy: Available")
                    else:
                        remedy_available = None
                        print(f"        → Remedy: Unknown")

                    recall_info['recalls'].append({
                        'number': recall_number,
                        'description': recall_description,
                        'remedy_available': remedy_available
                    })

                except Exception as e:
                    print(f"        ⚠ Error processing recall {idx+1}: {str(e)[:100]}")
                    continue
            
            # Compile results
            if recall_info['recalls']:
                recall_info['hasRecall'] = True
                print(f"  → Found {len(recall_info['recalls'])} recall(s)")
                for i, rec in enumerate(recall_info['recalls'], 1):
                    if rec['remedy_available'] is True:
                        remedy_status = "Available"
                    elif rec['remedy_available'] is False:
                        remedy_status = "NOT AVAILABLE"
                    else:
                        remedy_status = "Unknown"
                    print(f"      [{i}] {rec['number']}: {rec['description'][:50]}... (Remedy: {remedy_status})")
            
        except Exception as e:
            print(f"  ⚠ Error parsing recall details: {str(e)[:150]}")
            # Check again if it's really a "no recalls" case
            if 'no recalls' in body_text.lower() or 'there are no recalls' in body_text.lower():
                return {
                    'hasRecall': False,
                    'recalls': []
                }
            # If we truly can't tell, mark as error
            recall_info['hasRecall'] = None
            recall_info['recalls'] = [{'number': 'ERROR - Check manually', 'description': 'Error extracting details - check Ford website', 'remedy_available': None}]
        
        return recall_info
        
    except Exception as e:
        print(f"  ✗ Error: {str(e)[:150]}")
        return {
            'hasRecall': None,
            'recalls': [{'number': 'ERROR', 'description': f'Error: {str(e)[:150]}', 'remedy_available': None}]
        }

def process_recalls(input_txt_file, output_file):
    """
    Process VINs from text file and create Excel results file
    """
    print(f"Opening file: {input_txt_file}")
    
    # Read VINs from text file
    vins = []
    with open(input_txt_file, 'r') as f:
        for line in f:
            vin = line.strip()
            if vin:  # Skip empty lines
                vins.append(vin)
    
    if not vins:
        print("ERROR: No VINs found in text file")
        return
    
    print(f"Found {len(vins)} VINs in text file")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    results_sheet = wb.active
    results_sheet.title = 'RECALL_RESULTS'
    
    # Write headers - we'll start with base headers and expand as needed
    # Start with VIN, Recall status, then triplets of (Recall #, Description, Remedy Available) columns
    base_headers = ['VIN', 'Has Recall']
    results_sheet.append(base_headers)
    
    # Setup Selenium driver (start with Chrome)
    current_browser = 'chrome'
    print(f"\nInitializing {current_browser.upper()} browser...")
    driver = setup_driver(current_browser)

    # Setup debug log file
    output_dir = os.path.dirname(output_file)
    log_file = setup_debug_log(output_dir)
    print(f"Debug log: {log_file.name}")

    # Track the maximum number of recalls found on any vehicle
    max_recalls_found = 0
    
    # Store results temporarily so we can adjust headers after processing
    temp_results = []
    
    try:
        # Process ALL VINs
        total_vins = len(vins)
        processed = 0
        with_recalls = 0
        no_recalls = 0
        errors = 0
        
        print(f"\nProcessing {total_vins} VINs...\n")
        
        for idx, vin in enumerate(vins, 1):
            print(f"\n[{idx}/{total_vins}] Checking VIN: {vin}")
            
            # Check recall
            recall_data = check_ford_recall(driver, vin, log_file)
            
            # Track statistics
            if recall_data['hasRecall'] is None:
                errors += 1
            elif recall_data['hasRecall']:
                with_recalls += 1
            else:
                no_recalls += 1
            
            # ONLY ADD TO RESULTS IF THERE IS A RECALL
            if recall_data['hasRecall'] is True:
                # Track max recalls for header creation
                num_recalls = len(recall_data['recalls'])
                if num_recalls > max_recalls_found:
                    max_recalls_found = num_recalls
                
                # Store the result temporarily
                temp_results.append({
                    'vin': vin,
                    'recalls': recall_data['recalls']
                })
                
                print(f"  ✓ RECALL FOUND! ({num_recalls} recall(s))")
            elif recall_data['hasRecall'] is False:
                print(f"  ✓ No recalls (skipping)")
            else:
                print(f"  ⚠ Error checking (skipping)")
            
            processed += 1

            # Alternate browser every 40 VINs to reset session (Ford changes behavior after ~60)
            if idx % 40 == 0 and idx < total_vins:
                driver.quit()
                time.sleep(2)
                # Alternate between Chrome and Edge
                current_browser = 'edge' if current_browser == 'chrome' else 'chrome'
                print(f"\n  → Switching to {current_browser.upper()} browser...")
                driver = setup_driver(current_browser)
                print(f"  → Browser switched\n")

            # Small delay between requests
            time.sleep(1)
    
    finally:
        # Always close the driver and log file
        print("\nClosing browser...")
        driver.quit()
        if log_file:
            log_file.close()
            print(f"Debug log saved to: {log_file.name}")
    
    # Now create the proper headers based on max_recalls_found
    final_headers = ['VIN', 'Has Recall']
    for i in range(1, max_recalls_found + 1):
        final_headers.append(f'Recall #{i}: Number')
        final_headers.append(f'Recall #{i}: Description')
        final_headers.append(f'Recall #{i}: Remedy')
    
    # Clear the sheet and write proper headers
    results_sheet.delete_rows(1, 1)
    results_sheet.append(final_headers)
    
    # Now write all the results with proper column alignment
    for result in temp_results:
        row_data = [result['vin'], 'Yes']
        
        # Add each recall's number, description, and remedy availability
        for recall in result['recalls']:
            row_data.append(recall['number'])
            row_data.append(recall['description'])
            # Convert boolean to Yes/No, or keep None for errors
            if recall['remedy_available'] is True:
                row_data.append('Yes')
            elif recall['remedy_available'] is False:
                row_data.append('No')
            else:
                row_data.append('Unknown')
        
        # Fill remaining columns with "--" if this vehicle has fewer recalls than max
        while len(row_data) < len(final_headers):
            row_data.append('--')
        
        results_sheet.append(row_data)
    
    # Add left borders to all "Recall #X: Number" columns
    # Define a solid black border for the left side
    left_border = Border(left=Side(style='thin', color='000000'))
    
    # Find which columns are "Recall #X: Number" columns (they start at column 3, then every 3 columns)
    # Column indices: C=3, F=6, I=9, etc.
    for row_idx in range(1, results_sheet.max_row + 1):  # Include header row
        for recall_num in range(1, max_recalls_found + 1):
            # Calculate column index for "Recall #X: Number"
            # VIN=1, Has Recall=2, then each recall group has 3 columns
            col_idx = 2 + (recall_num - 1) * 3 + 1  # +1 for the Number column
            cell = results_sheet.cell(row=row_idx, column=col_idx)
            cell.border = left_border
    
    # Save the workbook
    wb.save(output_file)
    
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total Processed: {processed}")
    print(f"With Recalls: {with_recalls} (added to results)")
    print(f"No Recalls: {no_recalls} (not added to results)")
    print(f"Errors: {errors} (not added to results)")
    print(f"Max recalls on any vehicle: {max_recalls_found}")
    print(f"\nResults saved to: {output_file}")
    print(f"RECALL_RESULTS sheet contains ONLY vehicles with recalls ({with_recalls} vehicles)")
    print(f"Each recall has three columns: Recall #, Description, and Remedy Available")
    
    if errors > 0:
        print(f"\n⚠ WARNING: {errors} VIN(s) had errors.")

if __name__ == "__main__":
    import os
    
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Input file is VINS.txt in the same folder as the script
    input_file = os.path.join(script_dir, "VINS.txt")
    
    # Output directory is "Outputs" subfolder
    output_dir = os.path.join(script_dir, "Outputs")
    
    # Create Outputs folder if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}\n")
    
    # Create output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"FORD_RECALLS_{timestamp}.xlsx")
    
    print("="*60)
    print("FORD VIN RECALL CHECKER (Text File Version)")
    print("="*60)
    print()
    
    process_recalls(input_file, output_file)
