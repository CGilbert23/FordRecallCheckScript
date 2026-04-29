import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime
import re
import os
import json
import zipfile
import tempfile


def _build_proxy_auth_extension(host, port, user, password):
    """Build an in-memory Chrome extension that handles proxy auth.
    Chrome doesn't accept user:pass in --proxy-server, so we use the
    webRequest.onAuthRequired API to inject credentials at request time."""
    manifest = {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Proxy Auth",
        "permissions": [
            "proxy", "tabs", "unlimitedStorage", "storage",
            "webRequest", "webRequestBlocking", "<all_urls>"
        ],
        "background": {"scripts": ["background.js"]},
        "minimum_chrome_version": "22.0.0"
    }
    background_js = (
        'var config = {mode:"fixed_servers",rules:{singleProxy:{scheme:"http",'
        'host:%s,port:parseInt(%s)},bypassList:["localhost"]}};'
        'chrome.proxy.settings.set({value:config,scope:"regular"},function(){});'
        'chrome.webRequest.onAuthRequired.addListener('
        'function(details){return {authCredentials:{username:%s,password:%s}};},'
        '{urls:["<all_urls>"]},["blocking"]);'
    ) % (json.dumps(host), json.dumps(str(port)),
         json.dumps(user), json.dumps(password))

    fd, path = tempfile.mkstemp(suffix='.zip', prefix='proxy_auth_')
    os.close(fd)
    with zipfile.ZipFile(path, 'w') as zf:
        zf.writestr('manifest.json', json.dumps(manifest))
        zf.writestr('background.js', background_js)
    return path


def setup_driver():
    """Setup headless Chrome driver with anti-detection options"""
    from selenium.webdriver.chrome.service import Service
    import shutil
    import logging
    logger = logging.getLogger(__name__)

    proxy_host = os.environ.get('PROXY_HOST')
    proxy_port = os.environ.get('PROXY_PORT')
    proxy_user = os.environ.get('PROXY_USER')
    proxy_pass = os.environ.get('PROXY_PASS')
    use_proxy = all([proxy_host, proxy_port, proxy_user, proxy_pass])

    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--log-level=3')
    if not use_proxy:
        chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_argument('--remote-debugging-port=9222')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

    if use_proxy:
        ext_path = _build_proxy_auth_extension(proxy_host, proxy_port, proxy_user, proxy_pass)
        chrome_options.add_extension(ext_path)
        print(f"[SCRAPER] proxy enabled: {proxy_host}:{proxy_port} (user={proxy_user[:4]}***)", flush=True)
    else:
        print("[SCRAPER] no PROXY_* env vars set — direct connection (Ford will block from datacenter IPs)", flush=True)

    # Support custom Chrome binary (e.g. in Docker with Chromium)
    chrome_bin = os.environ.get('CHROME_BIN')
    if chrome_bin:
        chrome_options.binary_location = chrome_bin
        logger.info(f"Using Chrome binary: {chrome_bin}")

    # Find chromedriver
    chromedriver_path = shutil.which('chromedriver')
    if chromedriver_path:
        logger.info(f"Using chromedriver: {chromedriver_path}")
        service = Service(executable_path=chromedriver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
    else:
        logger.info("Using default chromedriver detection")
        driver = webdriver.Chrome(options=chrome_options)

    # Hide navigator.webdriver flag
    try:
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
        })
    except Exception as e:
        logger.warning(f"Could not set CDP command: {e}")

    return driver


def close_survey_popup(driver):
    """Close any survey popup that appears"""
    close_selectors = [
        'button[data-aut="button-close"]',
        'button[aria-label="Close"]',
        'button[aria-label="close"]',
        '.modal-close',
        'button.close',
        'div[class*="QSIWebResponsive"] button',
        'div[class*="modal"] button[class*="close"]',
        'div[class*="survey"] button',
    ]

    for selector in close_selectors:
        try:
            close_btns = driver.find_elements(By.CSS_SELECTOR, selector)
            for btn in close_btns:
                if btn.is_displayed():
                    try:
                        btn.click()
                        time.sleep(0.5)
                        return True
                    except:
                        pass
        except:
            pass

    try:
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        time.sleep(0.3)
    except:
        pass

    return False


def wait_for_overlays_to_clear(driver, timeout=3):
    """Wait for any overlays/modals to disappear before interacting with elements"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, '.modal-overlay, .popup-overlay, [class*="overlay"]:not([class*="no-overlay"])'))
        )
    except:
        pass
    close_survey_popup(driver)


def setup_debug_log(output_dir):
    """Create a debug log file for tracking navigation issues"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(output_dir, f"DEBUG_LOG_{timestamp}.txt")
    return open(log_path, 'w')


def debug_log(log_file, vin, message):
    """Write debug message to log file"""
    if log_file:
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_file.write(f"[{timestamp}] VIN {vin}: {message}\n")
        log_file.flush()


_diag_dump_count = 0
_DIAG_DUMP_LIMIT = 5


def diag_dump(driver, vin, reason, log_file):
    """Save page HTML + screenshot when scraping silently returns no recalls,
    so we can see what Ford actually returned. Capped at _DIAG_DUMP_LIMIT per run."""
    global _diag_dump_count
    if _diag_dump_count >= _DIAG_DUMP_LIMIT or not log_file:
        return
    _diag_dump_count += 1
    output_dir = os.path.dirname(log_file.name)
    timestamp = datetime.now().strftime("%H%M%S")
    base = os.path.join(output_dir, f"DIAG_{vin}_{timestamp}")
    try:
        with open(base + ".html", 'w', encoding='utf-8') as f:
            f.write(f"<!-- VIN: {vin} | reason: {reason} | url: {driver.current_url} -->\n")
            f.write(driver.page_source)
        driver.save_screenshot(base + ".png")
        debug_log(log_file, vin, f"DIAG dumped ({reason}): {base}.html + .png")
    except Exception as e:
        debug_log(log_file, vin, f"DIAG dump failed: {str(e)[:100]}")


def _live(msg):
    """Print to stdout so it shows up in `docker logs -f ford-checker`."""
    print(f"[SCRAPER] {msg}", flush=True)


def check_ford_recall(driver, vin, log_file=None):
    """
    Check Ford recall status for a given VIN using Selenium
    Returns: dict with hasRecall and recalls list
    """
    url = "https://www.ford.com/support/recalls-details/"
    _live(f"VIN {vin}: starting")

    try:
        wait = WebDriverWait(driver, 15)

        driver.get(url)
        _live(f"VIN {vin}: page loaded — title='{driver.title}' url={driver.current_url}")
        try:
            initial_body = driver.find_element(By.TAG_NAME, "body").text
            _live(f"VIN {vin}: body[0:400]={initial_body[:400]!r}")
        except Exception as e:
            _live(f"VIN {vin}: could not read body: {str(e)[:100]}")
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
        except Exception:
            _live(f"VIN {vin}: VIN INPUT NOT FOUND within 15s — selector '[data-testid=\"vin-search-text-field\"]' missing on page")
            try:
                inputs = driver.find_elements(By.TAG_NAME, "input")
                _live(f"VIN {vin}: page has {len(inputs)} <input> elements")
                for i, el in enumerate(inputs[:8]):
                    _live(f"VIN {vin}:   input[{i}] testid={el.get_attribute('data-testid')} name={el.get_attribute('name')} placeholder={el.get_attribute('placeholder')!r}")
            except Exception as e:
                _live(f"VIN {vin}: could not enumerate inputs: {str(e)[:100]}")
            diag_dump(driver, vin, "vin_input_not_found", log_file)
            raise
        debug_log(log_file, vin, f"URL after load: {driver.current_url}")

        wait_for_overlays_to_clear(driver)

        max_retries = 3
        for attempt in range(max_retries):
            try:
                close_survey_popup(driver)
                vin_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", vin_input)
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", vin_input)
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    debug_log(log_file, vin, f"Click attempt {attempt + 1} failed: {str(e)[:50]}")
                    time.sleep(1)
                    driver.execute_script("window.scrollTo(0, 0);")
                    wait_for_overlays_to_clear(driver)
                else:
                    raise
        time.sleep(0.3)

        debug_log(log_file, vin, f"Attempting to enter VIN: {vin}")

        before_value = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value BEFORE clear: '{before_value}'")

        vin_input.send_keys(Keys.CONTROL + "a")
        time.sleep(0.1)
        vin_input.send_keys(Keys.DELETE)
        time.sleep(0.1)

        after_clear = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value AFTER clear: '{after_clear}'")

        for char in vin:
            vin_input.send_keys(char)
        time.sleep(0.2)

        entered_value = vin_input.get_attribute('value')
        debug_log(log_file, vin, f"Field value AFTER typing: '{entered_value}'")

        if entered_value != vin:
            debug_log(log_file, vin, f"WARNING: VIN mismatch! Field has: '{entered_value}', Expected: '{vin}'")
            driver.get(url)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
            time.sleep(0.5)
            vin_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="vin-search-text-field"]')))
            driver.execute_script("arguments[0].click();", vin_input)
            time.sleep(0.2)
            for char in vin:
                vin_input.send_keys(char)
            time.sleep(0.2)
            entered_value = vin_input.get_attribute('value')
            debug_log(log_file, vin, f"After retry, field value: '{entered_value}'")

        debug_log(log_file, vin, f"VIN entered, about to submit")
        time.sleep(0.5)
        vin_input.send_keys(Keys.RETURN)
        debug_log(log_file, vin, f"RETURN pressed")

        time.sleep(2)

        try:
            page_text = driver.find_element(By.TAG_NAME, "body").text
            if "unexpected error" in page_text.lower() or "error occurred" in page_text.lower():
                debug_log(log_file, vin, f"ERROR MESSAGE DETECTED on page!")
                screenshot_path = os.path.join(os.path.dirname(log_file.name), f"ERROR_{vin}_{datetime.now().strftime('%H%M%S')}.png")
                driver.save_screenshot(screenshot_path)
                debug_log(log_file, vin, f"Screenshot saved: {screenshot_path}")
                error_elements = driver.find_elements(By.CSS_SELECTOR, '[class*="error"], [class*="Error"], [role="alert"]')
                for el in error_elements[:3]:
                    debug_log(log_file, vin, f"Error element text: {el.text[:100] if el.text else 'empty'}")
        except Exception as e:
            debug_log(log_file, vin, f"Error checking for errors: {str(e)[:50]}")

        time.sleep(3)

        debug_log(log_file, vin, f"URL after submit: {driver.current_url}")

        if '/recalls-details/' not in driver.current_url:
            _live(f"VIN {vin}: REDIRECT after submit -> {driver.current_url}")
            debug_log(log_file, vin, f"Redirect detected, navigating back...")
            diag_dump(driver, vin, "redirect_after_submit", log_file)
            driver.get(url)
            time.sleep(3)
            debug_log(log_file, vin, f"URL after redirect: {driver.current_url}")
        else:
            debug_log(log_file, vin, "No redirect, waiting for results...")
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: 'no recalls' in d.find_element(By.TAG_NAME, "body").text.lower() or
                              d.find_elements(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]') or
                              'there are no' in d.find_element(By.TAG_NAME, "body").text.lower()
                )
                debug_log(log_file, vin, "Results detected on page")
            except:
                debug_log(log_file, vin, "Timeout waiting for results, continuing anyway")
                time.sleep(2)

        body_text = driver.find_element(By.TAG_NAME, "body").text
        debug_log(log_file, vin, f"Page title: {driver.title}")
        _live(f"VIN {vin}: page title='{driver.title}' url={driver.current_url}")
        _live(f"VIN {vin}: body[0:300]={body_text[:300]!r}")

        if 'no recalls' in body_text.lower() or 'there are no recalls' in body_text.lower():
            _live(f"VIN {vin}: RESULT = no recalls (text match)")
            return {
                'hasRecall': False,
                'recalls': []
            }

        recall_info = {
            'hasRecall': False,
            'recalls': []
        }

        try:
            safety_header = driver.find_elements(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]')

            if not safety_header:
                _live(f"VIN {vin}: RESULT = no recalls (safety_header selector NOT FOUND — likely Ford changed markup)")
                diag_dump(driver, vin, "no_safety_header", log_file)
                return {
                    'hasRecall': False,
                    'recalls': []
                }

            try:
                parent = safety_header[0].find_element(By.XPATH, '..')
                tablist = parent.find_element(By.CSS_SELECTOR, '[role="tablist"]')
                recall_buttons = tablist.find_elements(By.CSS_SELECTOR, 'button[data-testid^="button-"][role="tab"]')
            except:
                recall_buttons = []

            if not recall_buttons:
                _live(f"VIN {vin}: RESULT = no recalls (safety_header found but no recall_buttons in tablist)")
                diag_dump(driver, vin, "no_recall_buttons", log_file)
                return {
                    'hasRecall': False,
                    'recalls': []
                }
            _live(f"VIN {vin}: found {len(recall_buttons)} recall button(s) — extracting")

            for idx in range(len(recall_buttons)):
                try:
                    safety_header = driver.find_element(By.CSS_SELECTOR, '[data-testid="button-safety-recalls-section-header"]')
                    parent = safety_header.find_element(By.XPATH, '..')
                    tablist = parent.find_element(By.CSS_SELECTOR, '[role="tablist"]')
                    buttons = tablist.find_elements(By.CSS_SELECTOR, 'button[data-testid^="button-"][role="tab"]')

                    if idx >= len(buttons):
                        break

                    button = buttons[idx]

                    testid = button.get_attribute('data-testid') or ''
                    recall_number = testid.replace('button-', '') if testid.startswith('button-') else None

                    if not recall_number:
                        continue

                    close_survey_popup(driver)

                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                    time.sleep(0.5)

                    recall_description = driver.execute_script("""
                        var p = arguments[0].querySelector('p');
                        return p ? p.innerText : '';
                    """, button) or 'See Ford website for details'

                    driver.execute_script("arguments[0].click();", button)
                    time.sleep(1.5)

                    try:
                        panel_id = f"content-panel-{recall_number}"
                        panel = driver.find_element(By.ID, panel_id)

                        campaign_text = driver.execute_script("""
                            var panel = arguments[0];
                            var sections = panel.querySelectorAll('section');
                            for (var i = 0; i < sections.length; i++) {
                                var label = sections[i].querySelector('p');
                                if (label && label.innerText.includes('Campaign')) {
                                    var valueP = sections[i].querySelectorAll('p')[1];
                                    if (valueP) return valueP.innerText.trim();
                                }
                            }
                            return null;
                        """, panel)

                        if campaign_text:
                            recall_number = campaign_text
                    except:
                        pass

                    page_text = driver.find_element(By.TAG_NAME, "body").text

                    if "Recall service not available right now" in page_text:
                        remedy_available = False
                    elif "Schedule your free recall service with a dealer" in page_text:
                        remedy_available = True
                    else:
                        remedy_available = None

                    recall_info['recalls'].append({
                        'number': recall_number,
                        'description': recall_description,
                        'remedy_available': remedy_available
                    })

                except Exception as e:
                    continue

            if recall_info['recalls']:
                recall_info['hasRecall'] = True
                _live(f"VIN {vin}: RESULT = {len(recall_info['recalls'])} recall(s) extracted")

        except Exception as e:
            if 'no recalls' in body_text.lower() or 'there are no recalls' in body_text.lower():
                return {
                    'hasRecall': False,
                    'recalls': []
                }
            recall_info['hasRecall'] = None
            recall_info['recalls'] = [{'number': 'ERROR - Check manually', 'description': 'Error extracting details - check Ford website', 'remedy_available': None}]

        return recall_info

    except Exception as e:
        _live(f"VIN {vin}: EXCEPTION {type(e).__name__}: {str(e)[:200]}")
        return {
            'hasRecall': None,
            'recalls': [{'number': 'ERROR', 'description': f'Error: {str(e)[:150]}', 'remedy_available': None}]
        }


def process_recalls(vins, output_file, progress_callback=None, vin_units=None):
    """
    Process a list of VINs and create Excel results file.
    progress_callback: optional function that receives a dict with status updates.
    Returns a summary dict.
    """
    if not vins:
        return {'error': 'No VINs provided'}

    global _diag_dump_count
    _diag_dump_count = 0

    output_dir = os.path.dirname(output_file)
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    results_sheet = wb.active
    results_sheet.title = 'RECALL_RESULTS'
    base_headers = ['VIN', 'Has Recall']
    results_sheet.append(base_headers)

    driver = setup_driver()
    log_file = setup_debug_log(output_dir)

    max_recalls_found = 0
    temp_results = []

    total_vins = len(vins)
    processed = 0
    with_recalls = 0
    no_recalls = 0
    errors = 0

    def report_progress(**kwargs):
        if progress_callback:
            progress_callback({
                'current': processed,
                'total': total_vins,
                'with_recalls': with_recalls,
                'no_recalls': no_recalls,
                'errors': errors,
                **kwargs
            })

    try:
        for idx, vin in enumerate(vins, 1):
            report_progress(status='processing', vin=vin)

            recall_data = check_ford_recall(driver, vin, log_file)

            if recall_data['hasRecall'] is None:
                errors += 1
            elif recall_data['hasRecall']:
                with_recalls += 1
            else:
                no_recalls += 1

            if recall_data['hasRecall'] is True:
                num_recalls = len(recall_data['recalls'])
                if num_recalls > max_recalls_found:
                    max_recalls_found = num_recalls

                temp_results.append({
                    'vin': vin,
                    'recalls': recall_data['recalls']
                })

            processed += 1
            report_progress(status='processing', vin=vin)

            # Restart headless Chrome every 40 VINs to reset session
            if idx % 40 == 0 and idx < total_vins:
                driver.quit()
                time.sleep(2)
                driver = setup_driver()

            time.sleep(1)

    finally:
        driver.quit()
        if log_file:
            log_file.close()

    # Build final Excel
    has_units = vin_units is not None

    final_headers = []
    if has_units:
        final_headers.append('Unit #')
    final_headers.extend(['VIN', 'Has Recall'])
    for i in range(1, max_recalls_found + 1):
        final_headers.append(f'Recall #{i}: Number')
        final_headers.append(f'Recall #{i}: Description')
        final_headers.append(f'Recall #{i}: Remedy')

    results_sheet.delete_rows(1, 1)
    results_sheet.append(final_headers)

    for result in temp_results:
        row_data = []
        if has_units:
            row_data.append(vin_units.get(result['vin'], ''))
        row_data.extend([result['vin'], 'Yes'])
        for recall in result['recalls']:
            row_data.append(recall['number'])
            row_data.append(recall['description'])
            if recall['remedy_available'] is True:
                row_data.append('Yes')
            elif recall['remedy_available'] is False:
                row_data.append('No')
            else:
                row_data.append('Unknown')
        while len(row_data) < len(final_headers):
            row_data.append('--')
        results_sheet.append(row_data)

    left_border = Border(left=Side(style='thin', color='000000'))
    base_cols = 3 if has_units else 2
    for row_idx in range(1, results_sheet.max_row + 1):
        for recall_num in range(1, max_recalls_found + 1):
            col_idx = base_cols + (recall_num - 1) * 3 + 1
            cell = results_sheet.cell(row=row_idx, column=col_idx)
            cell.border = left_border

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, results_sheet.max_column + 1):
        cell = results_sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    results_sheet.auto_filter.ref = results_sheet.dimensions

    wide_cols = {'E', 'H', 'K', 'O'} if has_units else {'D', 'G', 'J', 'N'}
    for col in range(1, results_sheet.max_column + 1):
        letter = results_sheet.cell(row=1, column=col).column_letter
        results_sheet.column_dimensions[letter].width = 35 if letter in wide_cols else 25

    wb.save(output_file)

    summary = {
        'processed': processed,
        'with_recalls': with_recalls,
        'no_recalls': no_recalls,
        'errors': errors,
        'max_recalls': max_recalls_found,
        'output_file': output_file,
    }

    report_progress(status='complete', **summary)
    return summary
